"""
PV现金流明细Excel日志生成器

用于生成Excel格式的PV现金流明细日志，每年一个sheet，包含：
1. 保单信息和精算假设
2. 利率曲线信息
3. 月度现金流投射明细
4. PV原材料值计算明细
5. 利率计算明细
"""

from decimal import Decimal
from datetime import date, datetime
from typing import Dict, List, Optional, Tuple
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os


class PVCashFlowExcelLogger:
    """PV现金流明细Excel日志生成器"""
    
    def __init__(self, policy_no: str, certi_no: Optional[str] = None, excel_file_path: Optional[str] = None):
        """
        初始化Excel日志生成器
        
        Args:
            policy_no: 保单号
            certi_no: 批单号（可选）
            excel_file_path: Excel文件路径，如果为None则自动生成
        """
        self.policy_no = policy_no
        self.certi_no = certi_no
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # 删除默认sheet
        
        # 样式定义
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        self.subheader_font = Font(bold=True, size=10)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.left_alignment = Alignment(horizontal='left', vertical='center')
        self.right_alignment = Alignment(horizontal='right', vertical='center')
        
        # 存储每年的数据
        self.year_data: Dict[int, Dict] = {}
        
        # 文件路径
        if excel_file_path is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            logs_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'logs')
            os.makedirs(logs_dir, exist_ok=True)
            
            # 文件名包含批单号
            certi_part = f"_{certi_no}" if certi_no else ""
            excel_file_path = os.path.join(logs_dir, f"pv_cashflow_detail_{policy_no}{certi_part}_{timestamp}.xlsx")
        
        self.excel_file_path = excel_file_path
    
    def add_year_data(
        self,
        year: int,
        policy_info: Dict,
        assumptions: Dict,
        rate_curves: Dict[str, pd.DataFrame],
        cash_flows: pd.DataFrame,
        pv_calculations: List[Dict],
        discount_factor_details: List[Dict]
    ):
        """
        添加一年的数据
        
        Args:
            year: 年份
            policy_info: 保单信息字典
            assumptions: 精算假设字典
            rate_curves: 利率曲线字典，key为曲线类型（如'locked', 'current'），value为DataFrame
            cash_flows: 现金流DataFrame，包含列：YYYYMM, Date_Obj, Premium, IACF, Claims, Expenses
            pv_calculations: PV计算明细列表，每个元素包含字段名、计算过程等
            discount_factor_details: 折现因子计算明细列表
        """
        self.year_data[year] = {
            'policy_info': policy_info,
            'assumptions': assumptions,
            'rate_curves': rate_curves,
            'cash_flows': cash_flows,
            'pv_calculations': pv_calculations,
            'discount_factor_details': discount_factor_details
        }
    
    def _format_decimal(self, value) -> str:
        """格式化Decimal值"""
        if isinstance(value, Decimal):
            if value == 0:
                return "0.00"
            precision = 6 if abs(value) < Decimal('10') else 2
            return f"{value:,.{precision}f}"
        return str(value) if value is not None else ""
    
    def _write_header(self, ws, row: int, title: str, col_span: int = 1):
        """写入标题行"""
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
        cell = ws.cell(row=row, column=1)
        cell.value = title
        cell.font = Font(bold=True, size=12)
        cell.fill = self.header_fill
        cell.font = self.header_font
        cell.alignment = self.center_alignment
        return row + 1
    
    def _write_subheader(self, ws, row: int, title: str, col_span: int = 1):
        """写入子标题行"""
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
        cell = ws.cell(row=row, column=1)
        cell.value = title
        cell.font = self.subheader_font
        cell.fill = self.subheader_fill
        cell.alignment = self.left_alignment
        return row + 1
    
    def _write_table(self, ws, row: int, headers: List[str], data: List[List], 
                     start_col: int = 1, auto_width: bool = True):
        """写入表格"""
        # 写入表头
        for col_idx, header in enumerate(headers, start=start_col):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = self.subheader_font
            cell.fill = self.subheader_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        current_row = row + 1
        
        # 写入数据
        for data_row in data:
            for col_idx, value in enumerate(data_row, start=start_col):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = self._format_decimal(value) if isinstance(value, (Decimal, float, int)) else str(value)
                cell.alignment = self.left_alignment if col_idx == start_col else self.right_alignment
                cell.border = self.border
            current_row += 1
        
        # 自动调整列宽
        if auto_width:
            for col_idx in range(start_col, start_col + len(headers)):
                col_letter = get_column_letter(col_idx)
                max_length = 0
                for r in range(row, current_row):
                    cell_value = ws.cell(row=r, column=col_idx).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                ws.column_dimensions[col_letter].width = min(max(max_length + 2, 10), 50)
        
        return current_row
    
    def _create_year_sheet(self, year: int, data: Dict):
        """为某一年创建sheet"""
        ws = self.wb.create_sheet(title=f"{year}年")
        current_row = 1
        
        # 1. 保单信息
        current_row = self._write_header(ws, current_row, f"{year}年 - 保单信息", col_span=2)
        current_row += 1
        
        policy_info = data['policy_info']
        info_data = [
            ["保单号", policy_info.get('policy_no', '')],
            ["批单号", policy_info.get('certi_no', '')],  # 添加批单号
            ["签单日期", policy_info.get('under_write_date', '')],
            ["起保日期", policy_info.get('start_date', '')],
            ["终保日期", policy_info.get('end_date', '')],
            ["保修结束日期", policy_info.get('warranty_end_date', '')],
            ["签单保费", self._format_decimal(policy_info.get('written_premium', 0))],
        ]
        current_row = self._write_table(ws, current_row, ["项目", "数值"], info_data)
        current_row += 2
        
        # 2. 精算假设
        current_row = self._write_header(ws, current_row, f"{year}年 - 精算假设", col_span=2)
        current_row += 1
        
        assumptions = data['assumptions']
        assump_data = [
            ["赔付率", self._format_decimal(assumptions.get('loss_ratio', 0))],
            ["间接理赔费用率", self._format_decimal(assumptions.get('claim_expense_ratio', 0))],
            ["维持费用率", self._format_decimal(assumptions.get('maintenance_expense_ratio', 0))],
            ["非金融风险调整率", self._format_decimal(assumptions.get('ra_ratio', 0))],
            ["获取费用率", self._format_decimal(assumptions.get('acquisition_expense_ratio', 0))],
        ]
        current_row = self._write_table(ws, current_row, ["项目", "数值"], assump_data)
        current_row += 2
        
        # 3. 利率曲线信息
        rate_curves = data['rate_curves']
        for curve_name, curve_df in rate_curves.items():
            curve_title = "锁定利率曲线" if curve_name == 'locked' else "期末利率曲线"
            current_row = self._write_header(ws, current_row, f"{year}年 - {curve_title}", col_span=4)
            current_row += 1
            
            # 准备利率曲线数据
            if not curve_df.empty and 'term_month' in curve_df.columns and 'forward_disrate_value' in curve_df.columns:
                # 计算累计折现因子
                curve_data = []
                cum_discount_factor = Decimal('1.0')
                for idx, row_data in curve_df.head(100).iterrows():  # 只显示前100个期限
                    term = int(row_data['term_month'])
                    rate = Decimal(str(row_data['forward_disrate_value']))
                    cum_discount_factor /= (Decimal('1.0') + rate)
                    curve_data.append([
                        term,
                        f"{rate:.6f}",
                        f"{cum_discount_factor:.6f}",
                        f"第{term}个月远期利率"
                    ])
                
                headers = ["期限（月）", "远期月利率", "折现因子（累计）", "说明"]
                current_row = self._write_table(ws, current_row, headers, curve_data)
            else:
                ws.cell(row=current_row, column=1).value = "利率曲线数据为空"
                current_row += 1
            
            current_row += 2
        
        # 4. 月度现金流投射明细
        current_row = self._write_header(ws, current_row, f"{year}年 - 月度现金流投射明细", col_span=8)
        current_row += 1
        
        cash_flows = data['cash_flows']
        if not cash_flows.empty:
            # 获取间接理赔费用率
            assumptions = data['assumptions']
            claim_expense_ratio = Decimal(str(assumptions.get('claim_expense_ratio', 0)))
            
            # 准备现金流数据
            cf_data = []
            for idx, row_data in cash_flows.iterrows():
                yyyymm = row_data.get('YYYYMM', '')
                date_obj = row_data.get('Date_Obj', '')
                if isinstance(date_obj, date):
                    date_str = date_obj.strftime('%Y-%m-%d')
                else:
                    date_str = str(date_obj)
                
                # 判断是否在风险期
                warranty_end_str = policy_info.get('warranty_end_date', '')
                if warranty_end_str and isinstance(date_obj, date):
                    try:
                        if isinstance(warranty_end_str, str):
                            warranty_end = datetime.strptime(warranty_end_str, '%Y-%m-%d').date()
                        else:
                            warranty_end = warranty_end_str
                        in_risk = date_obj >= warranty_end
                        risk_mark = "✅" if in_risk else "❌"
                    except:
                        risk_mark = ""
                else:
                    risk_mark = ""
                
                # 计算赔付流出（包含间接理赔费用）
                claims_base = Decimal(str(row_data.get('Claims', 0)))
                claims_with_expense = claims_base * (Decimal('1.0') + claim_expense_ratio)
                
                cf_data.append([
                    yyyymm,
                    date_str,
                    self._format_decimal(row_data.get('Premium', 0)),
                    self._format_decimal(row_data.get('IACF', 0)),
                    self._format_decimal(claims_with_expense),
                    self._format_decimal(row_data.get('Expenses', 0)),
                    risk_mark,
                    "风险期内" if risk_mark == "✅" else "保修期内"
                ])
            
            headers = ["年月", "日期", "保费流入", "IACF流出", "赔付流出", "维持费用流出", "风险期", "说明"]
            current_row = self._write_table(ws, current_row, headers, cf_data)
        else:
            ws.cell(row=current_row, column=1).value = "现金流数据为空"
            current_row += 1
        
        current_row += 2
        
        # 5. PV原材料值计算明细（简化版，只显示关键字段）
        current_row = self._write_header(ws, current_row, f"{year}年 - PV原材料值汇总", col_span=3)
        current_row += 1
        
        pv_calculations = data.get('pv_calculations', [])
        if pv_calculations:
            pv_data = []
            for pv_item in pv_calculations:
                pv_data.append([
                    pv_item.get('field_name', ''),
                    pv_item.get('description', ''),
                    self._format_decimal(pv_item.get('value', 0))
                ])
            
            headers = ["PV字段名", "中文描述", "数值"]
            current_row = self._write_table(ws, current_row, headers, pv_data)
        else:
            ws.cell(row=current_row, column=1).value = "PV计算数据为空"
            current_row += 1
        
        current_row += 2
        
        # 6. 折现因子计算示例（显示前几个关键计算）
        current_row = self._write_header(ws, current_row, f"{year}年 - 折现因子计算示例", col_span=7)
        current_row += 1
        
        discount_details = data.get('discount_factor_details', [])
        if discount_details:
            # 只显示前10个示例
            sample_details = discount_details[:10]
            detail_data = []
            for detail in sample_details:
                detail_data.append([
                    detail.get('cf_date', ''),
                    self._format_decimal(detail.get('cf_amount', 0)),
                    detail.get('months_from_base', ''),
                    detail.get('rate_used', ''),
                    detail.get('discount_factor', ''),
                    self._format_decimal(detail.get('present_value', 0)),
                    detail.get('note', '')
                ])
            
            headers = ["现金流日期", "现金流金额", "距离基准月数", "使用利率", "折现因子", "现值", "说明"]
            current_row = self._write_table(ws, current_row, headers, detail_data)
        else:
            ws.cell(row=current_row, column=1).value = "折现因子计算明细为空"
            current_row += 1
    
    def save(self):
        """保存Excel文件"""
        # 为每年创建sheet
        for year in sorted(self.year_data.keys()):
            self._create_year_sheet(year, self.year_data[year])
        
        # 保存文件
        self.wb.save(self.excel_file_path)
        return self.excel_file_path

